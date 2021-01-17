#!/usr/bin/python3
# coding=utf8
# GitLab 周代码提交详细统计

import shutil
import time

from openpyxl import load_workbook

from code_statistics import *

imgurl = "https://ae01.alicdn.com/kf/U950a6ec6d17d4bd7b37afe16959d7cbdO.jpg"
Template_Path = "template/周统计模板.xlsx"  # 模板路径
Save_Path = report_path + "/week-group.xlsx"  # 保存路径

date_list = []  # 日期列表
name_list = []  # Git用户名列表
daily_master_statistics = {}  # 每日详情数据
daily_dev_statistics = {}


# 获取本周日期列表
def get_week_list():
    global date_list
    # 获取当前日期
    today = datetime.date.today()
    # 本周第一天和最后一天
    this_week_start = today - datetime.timedelta(today.weekday())
    this_week_end = today + datetime.timedelta(6 - today.weekday())

    # 遍历上月每一天日期
    while this_week_start <= this_week_end:
        date_list.append(str(this_week_start))
        this_week_start += datetime.timedelta(1)


# 获取Git用户名列表
def get_user_list(date, dictionary):
    global name_list
    for name in dictionary[date].keys():
        if name in name_list:
            continue
        else:
            name_list.append(name)


# 生成每日统计数据
def get_gitlab_daily(start_time, end_time, date):
    global daily_dev_statistics, daily_master_statistics
    gitlab_info(start_time, end_time, "group")  # 群组仓库统计

    # 收集数据
    daily_master_statistics[date] = gitlab_statistics_data(1).copy()  # master
    print("master: ", daily_master_statistics)

    daily_dev_statistics[date] = gitlab_statistics_data(2).copy()  # dev
    print("dev: ", daily_dev_statistics)

    # 清空数据，准备下一轮循环
    info_master.clear()
    info_other.clear()

    # 收集Git用户名
    get_user_list(date, daily_master_statistics)
    get_user_list(date, daily_dev_statistics)


# 数据处理写入到excel文件(表面，项目分支名)
def wirte_excel(sheet, branch_type):
    wb = load_workbook(Save_Path)
    ws_master = wb[sheet]  # 访问指定工作表

    if branch_type == "master":
        temp = daily_master_statistics
    else:
        temp = daily_dev_statistics

    # 写入日期到表格
    i = 2  # 列
    for date in date_list:
        ws_master.cell(row=1, column=i, value=date + " (" + get_week_num(date) + ")")
        i += 4

    # 写入统计内容到表格
    j = 3  # 行
    for name in name_list:
        k = 2  # 列
        # 列写入姓名
        ws_master.cell(row=j, column=1, value=name)
        for date in date_list:
            # 如果有代码提交记录
            if name in temp[date].keys():
                ws_master.cell(row=j, column=k, value=temp[date][name][0])
                ws_master.cell(row=j, column=k + 1, value=temp[date][name][1])
                ws_master.cell(row=j, column=k + 2, value=temp[date][name][2])
                ws_master.cell(row=j, column=k + 3, value=temp[date][name][3])
            else:
                ws_master.cell(row=j, column=k, value=0)
                ws_master.cell(row=j, column=k + 1, value=0)
                ws_master.cell(row=j, column=k + 2, value=0)
                ws_master.cell(row=j, column=k + 3, value=0)
            # 下一列
            k += 4
        # 下一行
        j += 1
    wb.save(Save_Path)


if __name__ == "__main__":
    print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) + "  [周详细-群组仓]")
    get_week_list()  # 生成本周日期列表
    print(date_list)

    # 获取数据中。。。
    for date in date_list:
        time1 = date + " 00:00:00"  # 开始时间
        time2 = date + " 23:59:59"  # 结束时间
        print("\n", time1, time2)
        get_gitlab_daily(time1, time2, date)

    shutil.copy(Template_Path, Save_Path)  # 拷贝模板

    wirte_excel("master", "master")  # master
    wirte_excel("dev", "dev")  # dev

    dingding("http://localhost/week-group.xlsx", imgurl, "[周]详细统计仅群组仓 👉点击下载")  # 不需要推送可以注释掉
